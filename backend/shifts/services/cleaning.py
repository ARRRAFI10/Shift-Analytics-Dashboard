"""Ingestion and cleaning pipeline for the raw shift spreadsheet.

The pure transformation (`clean`) is deliberately free of any database access so
it can be unit-tested on small fixtures. `import_from_file` wraps it with reading
and idempotent persistence for the management command.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from django.conf import settings
from django.db import transaction

from openpyxl import load_workbook

RAW_COLUMNS = ["DAY_DATE", "START", "END", "HOURS", "REASON"]

CLEAN = "clean"
CORRECTED = "corrected"
QUARANTINED = "quarantined"
FLAGGED = "flagged"


@dataclass(frozen=True)
class CleaningConfig:
    hours_tolerance: float
    max_shift_hours: float

    @classmethod
    def from_settings(cls) -> "CleaningConfig":
        return cls(
            hours_tolerance=float(settings.HOURS_TOLERANCE),
            max_shift_hours=float(settings.MAX_SHIFT_HOURS),
        )


@dataclass
class CleanRecord:
    row_index: int
    day_date: date
    start: datetime
    end: datetime
    duration_hours: float
    stated_hours: float | None
    reason: str
    status: str  # CLEAN | CORRECTED


@dataclass
class Issue:
    row_index: int
    raw_data: dict[str, Any]
    issue_type: str
    description: str
    action_taken: str  # corrected | quarantined | flagged


@dataclass
class CleaningResult:
    records: list[CleanRecord] = field(default_factory=list)
    issues: list[Issue] = field(default_factory=list)


def parse_timestamp(value: Any) -> datetime | None:
    """Return a UTC-aware datetime, or None when the value is genuinely absent.

    Raises ValueError when a value is present but cannot be parsed, so callers can
    tell "missing" apart from "malformed".
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    # fromisoformat handles the trailing Z only once it's a numeric offset.
    dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    return date.fromisoformat(text)


def _jsonable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _raw_dict(cells: tuple[Any, ...]) -> dict[str, Any]:
    return {col: _jsonable(cells[i]) for i, col in enumerate(RAW_COLUMNS)}


def read_rows(path: str | Path) -> list[tuple[Any, ...]]:
    """Read the source sheet, returning the data rows (header excluded)."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    workbook.close()
    return rows


def clean(rows: list[tuple[Any, ...]], config: CleaningConfig) -> CleaningResult:
    result = CleaningResult()
    seen_signatures: set[tuple] = set()

    for offset, cells in enumerate(rows):
        # 1-based row index that lines up with the spreadsheet (header is row 1).
        row_index = offset + 2
        raw = _raw_dict(cells)
        day_raw, start_raw, end_raw, hours_raw, reason_raw = cells

        reason = str(reason_raw).strip() if reason_raw is not None else ""
        if not reason:
            reason = "Unknown"

        # Exact duplicates are caught up front so a repeated-but-otherwise-valid
        # row is reported as a duplicate rather than re-analysed.
        signature = tuple(_jsonable(c) for c in cells)
        if signature in seen_signatures:
            result.issues.append(
                Issue(
                    row_index,
                    raw,
                    "duplicate_row",
                    "Exact duplicate of an earlier row; only the first copy is kept.",
                    QUARANTINED,
                )
            )
            continue
        seen_signatures.add(signature)

        try:
            start = parse_timestamp(start_raw)
        except ValueError:
            result.issues.append(
                Issue(row_index, raw, "unparseable_timestamp",
                      f"START value '{start_raw}' could not be parsed.", QUARANTINED)
            )
            continue
        try:
            end = parse_timestamp(end_raw)
        except ValueError:
            result.issues.append(
                Issue(row_index, raw, "unparseable_timestamp",
                      f"END value '{end_raw}' could not be parsed.", QUARANTINED)
            )
            continue

        if start is None or end is None:
            missing = "START" if start is None else "END"
            result.issues.append(
                Issue(row_index, raw, "missing_timestamp",
                      f"{missing} is missing; duration cannot be derived.", QUARANTINED)
            )
            continue

        status = CLEAN
        notes: list[str] = []

        # DAY_DATE: recoverable from a valid START when it's malformed.
        try:
            day_date = parse_date(day_raw)
        except (ValueError, TypeError):
            day_date = None
        if day_date is None:
            day_date = start.date()
            status = CORRECTED
            note = f"DAY_DATE '{day_raw}' is invalid; derived {day_date} from START."
            notes.append(note)
            result.issues.append(
                Issue(row_index, raw, "invalid_day_date", note, CORRECTED)
            )

        if end <= start:
            result.issues.append(
                Issue(row_index, raw, "non_positive_duration",
                      "END is not after START; ordering is impossible.", QUARANTINED)
            )
            continue

        duration = (end - start).total_seconds() / 3600.0

        if duration > config.max_shift_hours:
            result.issues.append(
                Issue(row_index, raw, "implausible_duration",
                      f"Computed span {duration:.2f}h exceeds the "
                      f"{config.max_shift_hours:.0f}h limit; likely bad data.",
                      QUARANTINED)
            )
            continue

        stated = float(hours_raw) if hours_raw is not None else None
        if stated is not None and stated < 0:
            status = CORRECTED
            result.issues.append(
                Issue(row_index, raw, "negative_stated_hours",
                      f"Stated HOURS {stated} is negative; using computed "
                      f"{duration:.2f}h.", CORRECTED)
            )
        elif stated is not None and abs(stated - duration) > config.hours_tolerance:
            status = CORRECTED
            result.issues.append(
                Issue(row_index, raw, "hours_mismatch",
                      f"Stated HOURS {stated} differs from computed "
                      f"{duration:.2f}h beyond tolerance; using computed value.",
                      CORRECTED)
            )

        result.records.append(
            CleanRecord(row_index, day_date, start, end, round(duration, 4),
                        stated, reason, status)
        )

    _flag_overlaps(result)
    return result


def _flag_overlaps(result: CleaningResult) -> None:
    """Log same-day time overlaps as observations; records are left untouched.

    With no employee/machine identifier we cannot tell whether overlapping shifts
    are a data error or simply different resources, so we keep them and surface
    the overlap rather than dropping or merging (see README).
    """
    by_day: dict[date, list[CleanRecord]] = {}
    for record in result.records:
        by_day.setdefault(record.day_date, []).append(record)

    for day, records in by_day.items():
        ordered = sorted(records, key=lambda r: r.start)
        for i, current in enumerate(ordered):
            overlap_count = sum(
                1 for other in ordered
                if other is not current
                and other.start < current.end
                and current.start < other.end
            )
            if overlap_count:
                raw = {
                    "DAY_DATE": day.isoformat(),
                    "START": current.start.isoformat(),
                    "END": current.end.isoformat(),
                    "REASON": current.reason,
                }
                result.issues.append(
                    Issue(
                        current.row_index, raw, "same_day_overlap",
                        f"Overlaps {overlap_count} other shift(s) on {day}; kept "
                        f"as a valid record (resource is unknown).",
                        FLAGGED,
                    )
                )


def seed_categories() -> None:
    """Ensure the configured non-productive reasons exist and are marked as such."""
    from shifts.models import ActivityCategory

    for name in settings.NON_PRODUCTIVE_CATEGORIES:
        ActivityCategory.objects.update_or_create(
            name=name.strip(), defaults={"is_productive": False}
        )


def _category_for(name: str):
    """Resolve a category, creating unknown reasons as productive by default."""
    from shifts.models import ActivityCategory

    non_productive = {n.strip() for n in settings.NON_PRODUCTIVE_CATEGORIES}
    category, _ = ActivityCategory.objects.get_or_create(
        name=name, defaults={"is_productive": name not in non_productive}
    )
    return category


@transaction.atomic
def import_from_file(
    path: str | Path | None = None, *, source: str = "command"
) -> "ImportOutcome":
    """Run the full pipeline and persist it as a new dataset version.

    Each import creates its own Dataset and attaches its records and issues to it,
    so earlier imports remain viewable instead of being overwritten. The newest
    import becomes the active (default) dataset, and versions beyond the configured
    retention cap are pruned oldest-first.

    `path` may be a filesystem path or an uploaded file object; `source` records
    how the import was triggered (command vs. upload).
    """
    from shifts.models import Dataset, IngestionIssue, ShiftRecord

    source_obj = path if path is not None else _resolve_data_path()
    filename = _source_name(source_obj)
    rows = read_rows(source_obj)
    result = clean(rows, CleaningConfig.from_settings())

    seed_categories()

    actions: dict[str, int] = {}
    for issue in result.issues:
        actions[issue.action_taken] = actions.get(issue.action_taken, 0) + 1

    dataset = Dataset.objects.create(
        filename=filename,
        source=source,
        record_count=len(result.records),
        issue_count=len(result.issues),
        actions=actions,
    )

    for record in result.records:
        ShiftRecord.objects.create(
            dataset=dataset,
            day_date=record.day_date,
            start=record.start,
            end=record.end,
            duration_hours=record.duration_hours,
            stated_hours=record.stated_hours,
            category=_category_for(record.reason),
            status=record.status,
        )

    IngestionIssue.objects.bulk_create(
        IngestionIssue(
            dataset=dataset,
            row_index=issue.row_index,
            raw_data=issue.raw_data,
            issue_type=issue.issue_type,
            description=issue.description,
            action_taken=issue.action_taken,
        )
        for issue in result.issues
    )

    Dataset.objects.exclude(pk=dataset.pk).update(is_active=False)
    dataset.is_active = True
    dataset.save(update_fields=["is_active"])

    _enforce_retention(keep_active=dataset.pk)
    return ImportOutcome(dataset_id=dataset.pk, result=result)


@dataclass
class ImportOutcome:
    dataset_id: int
    result: CleaningResult


def _enforce_retention(keep_active: int) -> None:
    from shifts.models import Dataset

    cap = int(settings.MAX_RETAINED_DATASETS)
    if cap <= 0:
        return
    stale = (
        Dataset.objects.exclude(pk=keep_active)
        .order_by("-created_at", "-id")[cap - 1:]
    )
    Dataset.objects.filter(pk__in=[d.pk for d in stale]).delete()


def _resolve_data_path() -> Path:
    configured = Path(settings.DATA_FILE_PATH)
    return configured if configured.is_absolute() else settings.BASE_DIR / configured


def _source_name(source_obj: Any) -> str:
    if isinstance(source_obj, (str, Path)):
        return Path(source_obj).name
    # Uploaded files expose the original client filename via `.name`.
    return Path(str(getattr(source_obj, "name", "uploaded.xlsx"))).name
